# model.py
# Berisi semua logika data, pemrosesan AI, dan operasi inti.

import os
import json
from PIL import Image, ImageOps, ImageStat
from pdf2image import convert_from_path
import torch
import torch.nn as nn
from torchvision import transforms as T
from torchvision.models import mobilenet_v3_small
from torchvision.models.detection import ssdlite320_mobilenet_v3_large, SSDLite320_MobileNet_V3_Large_Weights 
from torchvision.models.detection.ssdlite import SSDLiteClassificationHead 
from torchvision.ops import nms
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
import numpy as np
import functools
import io
import config # Import file config

# Pastikan direktori sementara ada
os.makedirs(config.TEMP_PROCESSING_DIR, exist_ok=True)

# --- Fungsi Pemuatan Model ---
def load_char_classifier_model(model_path, num_classes, device_obj):
    if not os.path.exists(model_path):
        print(f"File model klasifikasi karakter tidak ditemukan di: {model_path}")
        return None
    try:
        model = mobilenet_v3_small(weights=None) 
        in_features_char = model.classifier[3].in_features
        model.classifier[3] = nn.Linear(in_features_char, num_classes)
        model.load_state_dict(torch.load(model_path, map_location=device_obj))
        model.to(device_obj)
        model.eval()
        print(f"Model klasifikasi karakter berhasil dimuat dari: {model_path}")
        return model
    except Exception as e:
        print(f"Gagal memuat model klasifikasi karakter: {e}")
        return None

def load_ssd_model(model_path, num_classes_ssd, device_obj):
    if not os.path.exists(model_path):
        print(f"File model SSD tidak ditemukan di: {model_path}")
        return None
    try:
        weights_ssd_coco = SSDLite320_MobileNet_V3_Large_Weights.COCO_V1
        model = ssdlite320_mobilenet_v3_large(weights=weights_ssd_coco) 
        in_channels_head_ssd = [672, 480, 512, 256, 256, 128] 
        num_anchors_ssd = model.anchor_generator.num_anchors_per_location()
        norm_layer_ssd = functools.partial(nn.BatchNorm2d, eps=0.001, momentum=0.03)
        model.head.classification_head = SSDLiteClassificationHead(
            in_channels=in_channels_head_ssd,
            num_anchors=num_anchors_ssd,
            num_classes=num_classes_ssd, 
            norm_layer=norm_layer_ssd
        )
        model.load_state_dict(torch.load(model_path, map_location=device_obj))
        model.to(device_obj)
        model.eval()
        print(f"Model SSD berhasil dimuat dari: {model_path}")
        return model
    except RuntimeError as e_runtime:
        print(f"Gagal memuat model SSD (RuntimeError): {e_runtime}")
        return None
    except Exception as e:
        print(f"Gagal memuat model SSD (error umum): {e}")
        return None

# --- Fungsi Transformasi ---
def get_char_transform(): # Diubah agar tidak menerima argumen
    return T.Compose([
        T.Resize(config.CHAR_IMAGE_SIZE),
        T.ToTensor(),
        T.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225]) 
    ])

def get_ssd_transform(): # Diubah agar tidak menerima argumen
    try:
        weights_enum_entry_ssd = SSDLite320_MobileNet_V3_Large_Weights.COCO_V1
        if callable(weights_enum_entry_ssd):
            weights_instance_ssd = weights_enum_entry_ssd()
        else: 
            weights_instance_ssd = weights_enum_entry_ssd
        if hasattr(weights_instance_ssd, 'transforms') and callable(weights_instance_ssd.transforms):
            print("Transformasi SSD menggunakan standar COCO pre-trained model (SSDLite320).")
            return weights_instance_ssd.transforms()
    except Exception as e_transform_ssd:
        print(f"Gagal mendapatkan transform dari SSDLite320_MobileNet_V3_Large_Weights: {e_transform_ssd}.")
    
    print("Menggunakan transformasi SSD fallback.")
    return T.Compose([
        T.Resize(config.SSD_INPUT_SIZE), 
        T.ToTensor(),
        T.ConvertImageDtype(torch.float) 
    ])

# --- Fungsi Logika Inti ---
def model_convert_pdf(pdf_bytes, dpi=300):
    temp_pdf_path = None
    try:
        temp_pdf_path = os.path.join(config.TEMP_PROCESSING_DIR, "temp_model_uploaded.pdf")
        with open(temp_pdf_path, "wb") as f:
            f.write(pdf_bytes)
        images = convert_from_path(temp_pdf_path, dpi=dpi, poppler_path=None)
        return images
    except Exception as e:
        print(f"Error saat model konversi PDF: {e}")
        return []
    finally: 
        if temp_pdf_path and os.path.exists(temp_pdf_path):
            try: os.remove(temp_pdf_path)
            except Exception as e_del: print(f"Gagal hapus temp PDF: {e_del}")

def model_crop_region(page_image_pil, annotation_value, original_width, original_height):
    img_w, img_h = page_image_pil.size
    scale_x = img_w / original_width
    scale_y = img_h / original_height
    x = annotation_value['x'] / 100 * original_width * scale_x
    y = annotation_value['y'] / 100 * original_height * scale_y
    w = annotation_value['width'] / 100 * original_width * scale_x
    h = annotation_value['height'] / 100 * original_height * scale_y
    return page_image_pil.crop((x, y, x + w, y + h))

def model_detect_chars(field_image_pil, model_ssd, device_obj, transform_ssd):
    if model_ssd is None: return []
    img_transformed = transform_ssd(field_image_pil.convert("RGB"))
    img_tensor = img_transformed.unsqueeze(0).to(device_obj) 
    model_ssd.eval()
    with torch.no_grad():
        predictions = model_ssd(img_tensor)
    pred = predictions[0]
    scores = pred['scores']
    labels = pred['labels']
    boxes = pred['boxes']
    target_label_mask = (labels == 1) & (scores > config.SSD_DETECTION_THRESHOLD)
    boxes_target_label = boxes[target_label_mask]
    scores_target_label = scores[target_label_mask]
    if boxes_target_label.nelement() == 0: return []
    keep_indices = nms(boxes_target_label, scores_target_label, config.SSD_NMS_IOU_THRESHOLD)
    nms_boxes = boxes_target_label[keep_indices].cpu().numpy().astype(int)
    
    final_char_boxes = []
    for box_coords in nms_boxes:
        x1, y1, x2, y2 = box_coords
        w_box = x2 - x1
        h_box = y2 - y1
        if w_box < config.MIN_CHAR_BOX_WIDTH or h_box < config.MIN_CHAR_BOX_HEIGHT or (w_box * h_box) < config.MIN_CHAR_AREA:
            continue
        aspect_ratio = w_box / h_box if h_box > 0 else float('inf')
        if aspect_ratio > config.MAX_CHAR_BOX_ASPECT_RATIO or \
           (1/aspect_ratio if aspect_ratio > 0 else float('inf')) > config.MAX_CHAR_BOX_ASPECT_RATIO:
            continue
        x1_c = max(0, x1)
        y1_c = max(0, y1)
        x2_c = min(field_image_pil.width, x2)
        y2_c = min(field_image_pil.height, y2)
        if x2_c <= x1_c or y2_c <= y1_c:
            continue
        final_char_boxes.append([x1_c, y1_c, x2_c, y2_c])

    if len(final_char_boxes) > 0:
        return sorted(final_char_boxes, key=lambda b: b[0])
    return []

def model_classify_char(char_image_pil, model_classifier, device_obj, transform_classifier):
    if model_classifier is None: return "?", 0.0
    if char_image_pil.width < config.MIN_CHAR_BOX_WIDTH or char_image_pil.height < config.MIN_CHAR_BOX_HEIGHT:
        return "?", 0.0
    try:
        gray_char = char_image_pil.convert('L')
        stat = ImageStat.Stat(gray_char)
        if stat.stddev[0] < 7 and stat.mean[0] > 240: 
            return "?", 0.0 
    except Exception: pass 
    img_tensor = transform_classifier(char_image_pil.convert("RGB")).unsqueeze(0).to(device_obj)
    model_classifier.eval()
    with torch.no_grad():
        output = model_classifier(img_tensor)
        probabilities = torch.softmax(output, dim=1)[0]
        confidence, predicted_idx = torch.max(probabilities, 0)
        predicted_char = config.CHAR_IDX_TO_CLASS.get(predicted_idx.item(), '?') 
    return predicted_char, confidence.item()

def create_excel_report(results_data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Hasil Ekstraksi"
    ws.append(["Nomor Pertanyaan", "Gambar Pertanyaan", "Hasil Karakter", "Avg. Confidence"]) 
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 45 
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20
    
    current_excel_row = 2
    temp_image_paths_for_cleanup = []

    for res_data in results_data:
        ws.row_dimensions[current_excel_row].height = 70 
        
        temp_img_path = os.path.join(config.TEMP_PROCESSING_DIR, f"excel_temp_img_{current_excel_row}.png")
        try:
            res_data["Image_PIL"].save(temp_img_path)
            temp_image_paths_for_cleanup.append(temp_img_path)
            
            img_for_excel = XLImage(temp_img_path)
            img_for_excel.height = 80 
            aspect_ratio = res_data["Image_PIL"].width / res_data["Image_PIL"].height if res_data["Image_PIL"].height > 0 else 1
            img_for_excel.width = img_for_excel.height * aspect_ratio
            ws.add_image(img_for_excel, f"B{current_excel_row}")
        except Exception as e_excel_img:
            print(f"Error menambah gambar ke excel: {e_excel_img}")
            ws[f"B{current_excel_row}"] = "Gagal memuat gambar"

        ws[f"A{current_excel_row}"] = res_data["ID_Pertanyaan"] 
        ws[f"C{current_excel_row}"] = res_data["Teks"]
        ws[f"D{current_excel_row}"] = res_data["Avg_Conf"]
        current_excel_row += 1
            
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    for img_path in temp_image_paths_for_cleanup:
        if os.path.exists(img_path):
            try: os.remove(img_path)
            except Exception as e_del_excel: print(f"Gagal hapus temp img excel: {e_del_excel}")
            
    return excel_buffer
